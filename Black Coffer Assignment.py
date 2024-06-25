#!/usr/bin/env python
# coding: utf-8

# # DATA ANALYSIS ASSIGNMENT

# 
# # Data Extraction

# In[1]:


import nltk
from nltk.tokenize import word_tokenize
from nltk.probability import FreqDist
from nltk.corpus import stopwords
import requests
from bs4 import BeautifulSoup
import pandas as pd


# In[2]:


nltk.download('punkt')
nltk.download('stopwords')


# In[3]:


def fetch_text_from_url(url):
    # Make an HTTP request to the URL
    response = requests.get(url)

    # Check if the request was successful (status code 200)
    if response.status_code == 200:
        # Parse the HTML content 
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Extract text
        text = soup.get_text()
        return text
    else:
        print(f"Error: Unable to fetch content from the URL. Status code: {response.status_code}")
        return None


# # Data Cleaning

# In[4]:


stopwords_files = ['StopWords_Auditor.txt', 'StopWords_Currencies.txt', 'StopWords_DatesandNumbers.txt' , 'StopWords_Generic.txt' , 'StopWords_GenericLong.txt', 'StopWords_Geographic.txt','StopWords_Names.txt']


# Read custom stopwords from each file
custom_stopwords = []

for filename in stopwords_files:
    with open(filename, 'r') as file:
        stopwords_list = [line.strip() for line in file]
        custom_stopwords.extend(stopwords_list)

def analyze_text(text):
    # Tokenize the text into words
    words = word_tokenize(text)

    # Remove stop words (common words that don't carry much meaning)
    stop_words = set(stopwords.words('english') + custom_stopwords)

    filtered_words = [word.lower() for word in words if word.isalpha() and word.lower() not in stop_words]
    x = len(filtered_words)
    

    # Calculate word frequency
    freq_dist = FreqDist(filtered_words)

    # Display the most common words
    print("Most common words:")
    for word, frequency in freq_dist.most_common(10):
        print(f"{word}: {frequency} times")


# In[5]:


df = pd.read_excel('Input.xlsx')
output_file = open('output_data.txt', 'w', encoding='utf-8')

# Iterate through the rows and perform textual analysis for each URL
for index, row in df.iterrows():
    url_id = row['URL_ID']
    url = row['URL']
    output_file.write(f"\nAnalyzing text for URL ID {url_id} ({url}):\n")


    
    print(f"\nAnalyzing text for URL ID {url_id} ({url}):")
    
    text_from_url = fetch_text_from_url(url)

    if text_from_url:
        analysis_result = analyze_text(text_from_url)
        if analysis_result is not None:
            output_file.write(analysis_result)
        


# # Data Cleansing

# In[6]:


with open('positive-words.txt','r') as file:
    positive_words = [word.strip() for word in file.readlines()]
print(positive_words)
with open('negative-words.txt','r') as file:
    negative_words = [word.strip() for word in file.readlines()]
print(negative_words)


# # Sentiment Analysis

# In[8]:



import requests
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.probability import FreqDist



def fetch_text_from_url(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an HTTPError for bad responses
        return response.text
    except requests.exceptions.HTTPError as err:
        print(f"Error: Unable to fetch content from the URL {url}. Status code: {err.response.status_code}")
        return None
k=1
from openpyxl import load_workbook

def analyze_sentiment(text, positive_words, negative_words,url_id,url,k):
    if text is None:
        print("Skipping sentiment analysis due to missing or invalid text.")
        return

    words = word_tokenize(text)
    stop_words = set(stopwords.words('english') + custom_stopwords)

    # Filter out stop words
    filtered_words = [word.lower() for word in words if word.isalpha() and word.lower() not in stop_words]
    total_words = len(filtered_words)

    # Create a frequency distribution
    freq_dist = FreqDist(filtered_words)
    print(freq_dist)

    # Calculate positive and negative scores
    positive_score = sum(freq_dist[word] for word in positive_words)
    negative_score = sum(freq_dist[word] for word in negative_words)

    sentiment = "Positive" if positive_score > negative_score else "Negative" if negative_score > positive_score else "Neutral"
    polarity_score = (positive_score-negative_score)/ ((positive_score + negative_score) + 0.000001)
    subjectivity_score = (positive_score+negative_score)/total_words + 0.000001
    
    new_data = {
    'url_id': url_id,
    'url': url,
    'positive_score': positive_score,
    'negative_score': negative_score,
    'polarity_score': polarity_score,
    'subjectivity_score': subjectivity_score}
    excel_file_path = 'sentiment_res.xlsx'

# Load the existing workbook
    try:
        wb = load_workbook(excel_file_path)
    except FileNotFoundError:
    # File doesn't exist, create a new workbook
        wb = Workbook()


    ws = wb.active

# Find the last row in the sheet
    last_row = ws.max_row + 1


    ws.cell(row=last_row, column=1, value=new_data['url_id'])
    ws.cell(row=last_row, column=2, value=new_data['url'])
    ws.cell(row=last_row, column=3, value=new_data['positive_score'])
    ws.cell(row=last_row, column=4, value=new_data['negative_score'])
    ws.cell(row=last_row, column=5, value=new_data['polarity_score'])
    ws.cell(row=last_row, column=6, value=new_data['subjectivity_score'])


    wb.save(excel_file_path)





    print(positive_score, negative_score, polarity_score,subjectivity_score, sentiment)

for index, row in df.iterrows():
    
    url_id = row['URL_ID']
    url = row['URL']
    output_file.write(f"\nAnalyzing text for URL ID {url_id} ({url}):\n")

    print(f"\nAnalyzing text for URL ID {url_id} ({url}):")
    
    text_from_url = fetch_text_from_url(url)
    analyze_sentiment(text_from_url, positive_words, negative_words,url_id,url,k)
    k+=1


# In[9]:


from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
from nltk.corpus import cmudict  # This is used to count syllables
import string
import re


import nltk
nltk.download('cmudict')
nltk.download('stopwords')


# In[10]:


def count_syllables(word, pronunciations):
    # Check if the word is in the pronunciation dictionary
    if word.lower() in pronunciations:
        return max([len(list(y for y in x if y[-1].isdigit())) for x in pronunciations[word.lower()]] or [0])
    else:
        return 0


# In[11]:


def count_personal_pronouns(text):
    personal_pronouns = set(['i', 'me', 'my', 'mine', 'myself', 'you', 'your', 'yours', 'yourself', 'he', 'him', 'his', 'himself',
                             'she', 'her', 'hers', 'herself', 'it', 'its', 'itself', 'we', 'us', 'our', 'ours', 'ourselves',
                             'they', 'them', 'their', 'theirs', 'themselves'])

    words = word_tokenize(text)
    personal_pronoun_count = sum(1 for word in words if word.lower() in personal_pronouns)
    return personal_pronoun_count


# In[12]:



def calculate_percentage_of_complex_words(text, pronunciations):
    words = word_tokenize(text)
    total_words = len(words)
    
    # Calculate syllables for each word
    syllables = [count_syllables(word, pronunciations) for word in words]
    
    # Count words with more than 2 syllables as complex words
    complex_word_count = sum(1 for syllable_count in syllables if syllable_count > 2)

    if total_words == 0:
        return 0

    # Calculate the percentage of complex words
    complex_word_percentage = (complex_word_count / total_words) * 100

    return complex_word_percentage


# In[13]:


def complex_words(text, pronunciations):
    words = word_tokenize(text)
    total_words = len(words)
    
    # Calculate syllables for each word
    syllables = [count_syllables(word, pronunciations) for word in words]
    
    # Count words with more than 2 syllables as complex words
    complex_word_count = sum(1 for syllable_count in syllables if syllable_count > 2)
    return complex_word_count


# In[14]:


def calculate_avg_words_per_sentence(text):
    sentences = sent_tokenize(text)
    total_words = sum(len(word_tokenize(sentence)) for sentence in sentences)
    total_sentences = len(sentences)

    if total_sentences == 0:
        return 0

    return total_words / total_sentences


# In[15]:


def remove_stopwords_and_punctuations(text):
    # Tokenize the text
    stop_words = set(stopwords.words('english'))

    # Remove stopwords and punctuations
    words = [word.lower() for word in word_tokenize(text) if word.isalpha() and word.lower() not in stop_words and word not in string.punctuation]

    return words


# In[16]:


def calculate_average_word_length(text):
    words = remove_stopwords_and_punctuations(text)
    total_word_length = sum(len(word) for word in words)
    total_words = len(words)

    if total_words == 0:
        return 0

    return total_word_length / total_words


# In[17]:


k = 2

def analyse_sentiment(text, positive_words, negative_words,url_id,url,last_row):
    if text is None:
        print("Skipping sentiment analysis due to missing or invalid text.")
        return

    cleaned_words = remove_stopwords_and_punctuations(text)
    
    # Calculate the Word Count
    word_count = len(cleaned_words)

    # Get the pronunciation dictionary for counting syllables
    pronunciations = cmudict.dict()

    # Calculate Average Sentence Length
    avg_sentence_length = calculate_avg_words_per_sentence(text)
    
    # Calculate Average Number of Words per Sentence
    avg_words_per_sentence = calculate_avg_words_per_sentence(text)

    # Calculate Percentage of Complex Words (words with more than 2 syllables)
    complex_word_count_percent = calculate_percentage_of_complex_words(text, pronunciations)

    # Calculate Fog Index
    fog_index = 0.4 * (avg_sentence_length + complex_word_count_percent)
    
    complex_word_count = complex_words(text, pronunciations)
    # Count Syllables Per Word
    syllable_count_per_word = [count_syllables(word, pronunciations) for word in cleaned_words]

    # Count Personal Pronouns
    personal_pronoun_count = count_personal_pronouns(text)

    # Calculate Average Word Length
    avg_word_length = calculate_average_word_length(text)

    print(f"Average Sentence Length: {avg_sentence_length}")
    print(f"Percentage of Complex Words: {complex_word_count_percent}%")
    print(f"Fog Index: {fog_index}")
    print(f"Average Number of Words per Sentence: {avg_words_per_sentence}")
    print(f"Word Count (after removing stopwords and punctuations): {word_count}")
    print(f"Syllable Count Per Word: {syllable_count_per_word}")
    print(f"Personal Pronoun Count: {personal_pronoun_count}")
    print(f"Average Word Length: {avg_word_length}")
    
    dt = {
        'URL_ID': url_id,
        'URL': url,
        'AVG SENTENCE LENGTH': avg_sentence_length,
        'COMPLEX WORD COUNT': complex_word_count,
        'complex_word_count_percent': complex_word_count_percent ,
        'FOG INDEX': fog_index,
        'AVG NUMBER OF WORDS PER SENTENCE': avg_words_per_sentence,
        'WORD COUNT': word_count,
        'SYLLABLE PER WORD': syllable_count_per_word,
        'PERSONAL PRONOUNS': personal_pronoun_count,
        'AVG WORD LENGTH': avg_word_length
    }
    excel_file_path = 'sentiment_res.xlsx'
    try:
        wb = load_workbook(excel_file_path)
    except FileNotFoundError:
    # File doesn't exist, create a new workbook
        wb = Workbook()


    ws = wb.active
    new_fields = ['AVG SENTENCE LENGTH', 'COMPLEX WORD COUNT', 'FOG INDEX','AVG NUMBER OF WORDS PER SENTENCE','WORD COUNT',
                 'SYLLABLE PER WORD','PERSONAL PRONOUNS','AVG WORD LENGTH','complex_word_count_percent']

# Determine the column index to start adding new fields
    start_column = ws.max_column + 1
    

# Insert new field names into the first row
    for idx, field in enumerate(new_fields, start=start_column):
        ws.cell(row=1, column=idx, value=field)



# Find the last row in the sheet
   
    
    ws.cell(row=last_row, column=7, value=dt['AVG SENTENCE LENGTH'])
    ws.cell(row=last_row, column=8, value=dt['COMPLEX WORD COUNT'])
    ws.cell(row=last_row, column=9, value=dt['FOG INDEX'])
    ws.cell(row=last_row, column=10, value=dt['AVG NUMBER OF WORDS PER SENTENCE'])
    ws.cell(row=last_row, column=11, value=dt['WORD COUNT'])
   
    ws.cell(row=last_row, column=13, value=dt['PERSONAL PRONOUNS'])
    ws.cell(row=last_row, column=14, value=dt['AVG WORD LENGTH'])
    ws.cell(row=last_row, column=15, value=dt['complex_word_count_percent'])
    last_row+=1
    wb.save(excel_file_path)
af = pd.read_excel('Output Data Structure.xlsx')
for index, row in af.iterrows():
    url_id = row['URL_ID']
    url = row['URL']
    output_file.write(f"\nAnalyzing text for URL ID {url_id} ({url}):\n")

    print(f"\nAnalyzing text for URL ID {url_id} ({url}):")
    
    text_from_url = fetch_text_from_url(url)
    
    k+=1
    result = analyse_sentiment(text_from_url, positive_words, negative_words,url_id,url,k)


# In[ ]:




